import openpyxl


def new_tab(wb, tab_name):
    """
    Creates and returns new xls tab
    """
    ws = wb.create_sheet()
    ws.title = tab_name

    return ws


def output_data_ss(ws, exp_data):
    """
    Adds single substrate data to current worksheet.
    """
    row = []
    for i in range(exp_data.replicates):
        row += ['Replicate {}'.format(i+1), '']
    ws.append(row)

    # naming row
    row = ['Concentration [{}]'.format(exp_data.cunit),
           'Rate (experimental) [{}]'.format(exp_data.runit)] * exp_data.replicates
    ws.append(row)

    # data rows
    row_nums = max([len(rep) for rep in exp_data.concentrations])
    for drow in range(row_nums):
        row = []
        for i in range(exp_data.replicates):
            if len(exp_data.rates[i]) - 1 < drow:
                row += ['', '']
                continue
            row += [exp_data.concentrations[i][drow],
                    exp_data.rates[i][drow]]
        ws.append(row)


def output_ds_subgroup(ws, subgroup, exp_data):
    varsets = exp_data.AllVar[subgroup]
    conssets = exp_data.AllConst[subgroup]
    ratesets = exp_data.AllRates[subgroup]
    set_nu = 0
    # iter over sets
    for varset, consset,  rateset in zip(varsets, conssets, ratesets):
        set_nu += 1
        # set naming
        ws.append(['Set {}'.format(set_nu)])
        row = []
        #  rep row
        for i in range(len(varset)):
            row += ['Replicate {}'.format(i+1), '', '']
        ws.append(row)
        # naming row
        row = ['Concentration ({}) [{}]'.format(exp_data.nameA if subgroup else exp_data.nameB, exp_data.cunit),
               'Concentration ({}) [{}]'.format(exp_data.nameB if subgroup else exp_data.nameA, exp_data.cunit),
               'Rate (experimental) [{}]'.format(exp_data.runit)] * len(varset)
        ws.append(row)

        # insert values
        for p_count in range(len(varset[0])):
            row = []
            for rep_count in range(len(varset)):
                row += [varset[rep_count][p_count], consset[rep_count][p_count], rateset[rep_count][p_count]]
            ws.append(row)
        ws.append([])


def output_data_ds(wb, exp_data):
    ws = new_tab(wb, exp_data.nameA + " - input_data")
    output_ds_subgroup(ws, True, exp_data)
    ws = new_tab(wb, exp_data.nameB + " - input_data")
    output_ds_subgroup(ws, False, exp_data)


def data_to_xls(data, w_file):
    """
    Write data to excel File
    """
    wb = openpyxl.Workbook()
    if data.is_single():
        ws = new_tab(wb, data.name + " - input_data")
        output_data_ss(ws, data)
    else:
        output_data_ds(wb, data)
    wb.worksheets = wb.worksheets[1:]
    wb.save(w_file)


def output_fit_ss(ws, exp_data, fit):
    """
    Adds single substrate data to current worksheet.
    """
    row = []
    for i in range(exp_data.replicates):
        row += ['Replicate {}'.format(i+1), '', '']
    ws.append(row)

    # naming row
    row = ['Concentration [{}]'.format(exp_data.cunit),
           'Rate (experimental) [{}]'.format(exp_data.runit),
           'Rate (predicted) [{}]'.format(exp_data.runit)] * exp_data.replicates
    ws.append(row)

    # data rows
    row_nums = max([len(rep) for rep in exp_data.concentrations])
    for drow in range(row_nums):
        row = []
        for i in range(exp_data.replicates):
            if len(exp_data.rates[i]) - 1 < drow:
                row += ['', '', '']
                continue
            row += [exp_data.concentrations[i][drow],
                    exp_data.rates[i][drow],
                    fit.function(exp_data.concentrations[i][drow])]
        ws.append(row)
    ws.append([])
    ws.append(['Fit parameters'])
    ws.append(['', 'Fitted value', 'Lower bound', 'Upper bound'])
    for param in fit.params:
        vals = fit.params[param]
        min_val = str(vals.min) if vals.min == float('-inf') else vals.min
        max_val = str(vals.max) if vals.max == float('-inf') else vals.max
        ws.append([param, vals.value, min_val, max_val, fit.get_units(param, exp_data)])


def output_fit_ds(wb, exp_data, fits):
    ws = new_tab(wb, exp_data.nameA + " - fitted_data")
    output_fit_subgroup(ws, exp_data, fits, True)
    ws = new_tab(wb, exp_data.nameB + " - fitted_data")
    output_fit_subgroup(ws, exp_data, fits, False)


def output_fit_subgroup(ws, exp_data, fits, subgroup):
    """
    Adds double substrate data to current worksheet.
    """
    varsets = exp_data.AllVar[subgroup]
    conssets = exp_data.AllConst[subgroup]
    ratesets = exp_data.AllRates[subgroup]
    enzymesets = exp_data.Econc[subgroup]

    set_nu = 0
    # iter over sets
    for varset, consset, rateset, enzymeset, fit in zip(varsets, conssets, ratesets, enzymesets, fits[subgroup]):
        set_nu += 1
        # set naming
        ws.append(['Set {}'.format(set_nu)])
        row = []
        #  rep row
        for i in range(len(varset)):
            row += ['Replicate {}'.format(i+1), '', '', '']
        ws.append(row)
        # naming row
        row = ['Concentration ({}) [{}]'.format(exp_data.nameA if subgroup else exp_data.nameB, exp_data.cunit),
               'Concentration ({}) [{}]'.format(exp_data.nameB if subgroup else exp_data.nameA, exp_data.cunit),
               'Rate (experimental) [{}]'.format(exp_data.runit),
               'Rate (predicted) [{}]'.format(exp_data.runit)] * len(varset)
        ws.append(row)
        # insert values
        for p_count in range(len(varset[0])):
            row = []
            for rep_count in range(len(varset)):
                row += [varset[rep_count][p_count], consset[rep_count][p_count], rateset[rep_count][p_count],
                        fit.function(varset[rep_count][p_count], consset[rep_count][p_count],
                                     enzymeset[rep_count][p_count])]
            ws.append(row)
        ws.append([])
        ws.append(['Fit parameters'])
        ws.append(['', 'Fitted value', 'Lower bound', 'Upper bound'])
        for param in fit.params:
            vals = fit.params[param]
            min_val = str(vals.min) if vals.min == float('-inf') else vals.min
            max_val = str(vals.max) if vals.max == float('-inf') else vals.max
            ws.append([param, vals.value, min_val, max_val, fit.get_units(param, exp_data)])
        ws.append([])
        ws.append([])


def fit_to_xls(data, fit, w_file):
    """
    Writes fitting results to excel file
    """
    wb = openpyxl.Workbook()
    if data.is_single():
        ws = new_tab(wb, "fit results")
        output_fit_ss(ws, data, fit)
    else:
        output_fit_ds(wb, data, fit)
    wb.worksheets = wb.worksheets[1:]
    wb.save(w_file)
